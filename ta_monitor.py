#!/usr/bin/env python3
"""
TA Monitor — Monitoraggio Posizioni HR / Talent Acquisition Roma
Maieutike srl — v2.1 (GitHub Actions edition)

Uso:
    python ta_monitor.py                              # Default: ultime 72h
    python ta_monitor.py --hours-old 168              # Ultima settimana
    python ta_monitor.py --search-terms "HR manager"  # Keywords custom
"""

import argparse
import csv
import hashlib
import json
import logging
import os
import re
import sqlite3
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import yaml
from rapidfuzz import fuzz

# === PATHS ===
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data"
OUTPUT_DIR = BASE_DIR / "output"
LOG_DIR = BASE_DIR / "logs"

for d in [DATA_DIR, OUTPUT_DIR, LOG_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ============================================================
# FILTRO DI RILEVANZA
# Carica da filters.json se presente (gestito dalla dashboard)
# Altrimenti usa i valori hardcoded sotto
# ============================================================

def _load_filters_json():
    """Carica filtri da filters.json se presente."""
    fpath = BASE_DIR / "filters.json"
    if fpath.exists():
        try:
            with open(fpath, "r", encoding="utf-8") as f:
                data = json.load(f)
            print(f"  [filters.json] Caricato v{data.get('_version', '?')}")
            return data
        except Exception as e:
            print(f"  [filters.json] Errore: {e} — uso filtri hardcoded")
    return None

_EXTERNAL_FILTERS = _load_filters_json()

def _get_patterns(key, default):
    if _EXTERNAL_FILTERS and key in _EXTERNAL_FILTERS:
        items = _EXTERNAL_FILTERS[key]
        if items and isinstance(items[0], dict):
            return [i["pattern"] for i in items]
        return items
    return default

def _get_agencies():
    if _EXTERNAL_FILTERS and "agencies" in _EXTERNAL_FILTERS:
        return _EXTERNAL_FILTERS["agencies"]
    return _DEFAULT_AGENCIES

_DEFAULT_AGENCIES = [
    "QuoJobis", "Gi Group", "Adecco", "Randstad", "Manpower",
    "Synergie", "OpenJobMetis", "LavoroPiù", "Etjca", "Umana",
    "MAW", "Orienta", "Humangest", "In-HR", "Tempor",
    "Generazione Vincente", "Ali Lavoro", "E-Work", "Eurointerim",
    "Kelly", "Hays", "Page Group", "Michael Page", "Spring",
    "Hunters", "Articolo1",
]

RELEVANT_TITLE_PATTERNS = [
    # Talent Acquisition
    r"talent.?acquisition",
    # HR Business Partner
    r"hr\s*b\.?p\.?",
    r"human\s*resource.*business\s*partner",
    r"hr\s+business\s+partner",
    # HR Generalist
    r"hr\s+generalist",
    r"human\s*resource.*generalist",
    # HR Manager / Director
    r"hr\s+manager",
    r"human\s*resource.*manager",
    r"dirett.*risorse\s+umane",
    r"responsabile\s+risorse\s+umane",
    r"hr\s+director",
    # Recruiting / Recruiter / Selezionatrice/ore
    r"recruit(?:er|ing|ment)",
    r"selezionat(?:ore|rice|rici)",
    r"selezione\s+(?:del\s+)?personale",
    r"responsabile\s+selezione",
    r"addett[oa]\s+(?:alla?\s+)?selezione",
    r"specialista\s+(?:della?\s+)?selezione",
    r"ricerca\s+e\s+selezione",
    # HR Specialist
    r"hr\s+specialist",
    r"specialist[a]?\s+(?:hr|risorse\s+umane)",
    r"hr\s+officer",
    r"people\s+(?:partner|manager|specialist|officer)",
    # HR Administration con focus selezione
    r"hr\s+admin.*selezione",
    # Head hunter / Executive search
    r"head\s*hunt(?:er|ing)",
    r"executive\s+search",
    # Employer Branding (spesso combinato con TA)
    r"employer\s+brand",
    # Gestione filiali agenzie per il lavoro / somministrazione
    # Pattern generici con contesto agenzia
    r"respons.*filiale.*(?:somministr|lavoro|apl|interinale|staffing)",
    r"branch\s*manager.*(?:staffing|agenzia|somministr|lavoro|apl)",
    r"dirett.*filiale.*(?:somministr|lavoro|apl|interinale|staffing)",
    r"gestione\s+filiale.*(?:somministr|lavoro|apl|interinale|staffing)",
    r"area\s+manager.*(?:somministr|staffing|apl|agenzia)",
    # Agenzie specifiche — qualsiasi ruolo di filiale
    r"(?:quojobis|gi\s*group|adecco|randstad|manpower|synergie|openjob|lavoropiu|etjca|umana|maw|orienta|humangest|in-hr|tempor|generazione\s*vincente|ali\s+lavoro|e-work|eurointerim|kelly|hays|page\s*group|michael\s*page|spring|hunters|articolo1).*(?:filiale|branch|responsabil|dirett|manager|account|hr\s+consultant)",
    r"(?:filiale|branch|responsabil|account).*(?:quojobis|gi\s*group|adecco|randstad|manpower|synergie|openjob|lavoropiu|etjca|umana|maw|orienta|humangest|in-hr|tempor|generazione\s*vincente|ali\s+lavoro|e-work|eurointerim|kelly|hays)",
    # Ruoli operativi in agenzie per il lavoro
    r"(?:account|consultant).*(?:apl|somministr|staffing|agenzia.*lavoro)",
    r"hr\s+consultant",
    r"staffing\s+(?:specialist|manager|consultant)",
]

_RELEVANT_COMPILED = [re.compile(p, re.IGNORECASE) for p in _get_patterns("relevant_patterns", RELEVANT_TITLE_PATTERNS)]

# Genera pattern dinamici per le agenzie caricate
def _build_agency_patterns():
    agencies = _get_agencies()
    if not agencies:
        return []
    # Normalizza nomi per regex: spazi → \s*, caratteri speciali escaped
    parts = []
    for a in agencies:
        # "Gi Group" → "gi\s*group", "Page Group" → "page\s*group"
        escaped = re.escape(a).replace(r"\ ", r"\s*")
        parts.append(escaped)
    agency_re = "|".join(parts)
    roles_re = r"(?:filiale|branch|responsabil|dirett|manager|account|hr\s+consultant|recruiter|specialist)"
    return [
        re.compile(f"(?:{agency_re}).*{roles_re}", re.IGNORECASE),
        re.compile(f"{roles_re}.*(?:{agency_re})", re.IGNORECASE),
    ]

_AGENCY_PATTERNS = _build_agency_patterns()

BROAD_KEYWORDS = _get_patterns("broad_keywords", [
    "talent acquisition", "talent", "recruiter", "recruiting", "recruitment",
    "selezione", "selezionatrice", "selezionatore",
    "hrbp", "hr generalist", "hr specialist", "hr manager", "hr officer",
    "people partner", "people manager",
    "head hunter", "headhunter", "executive search",
    "branch manager staffing", "branch manager agenzia",
    "responsabile filiale agenzia", "responsabile filiale apl",
    "risorse umane", "human resources",
    "hr consultant", "employer branding", "staffing specialist",
    "ricerca e selezione",
    # Agenzie specifiche (se appaiono nel titolo con ruoli HR)
    "quojobis", "gi group", "synergie", "openjobmetis", "lavoropiu",
    "etjca", "humangest", "umana", "orienta", "maw",
])

# Blacklist: titoli NON pertinenti
TITLE_BLACKLIST = _get_patterns("blacklist", [
    # Tech / IT
    r"sviluppat", r"developer", r"software\s+engineer", r"engineer(?:ing)?",
    r"full\s*stack", r"front\s*end", r"back\s*end", r"mobile\s+engineer",
    r"data\s+(?:engineer|scientist|analyst)", r"data\s+&\s+analytics",
    r"devops", r"sys\s*admin", r"cloud.*(?:architect|engineer|advisory|consultant)",
    r"network\s+engineer", r"system\s+engineer", r"UX\s+engineer",
    r"user\s+experience", r"AI\/?Gen\s*AI", r"artificial\s+intelligen",
    r"machine\s+learning", r"cyber\s*secur", r"infrastructure",
    r"murex\s+consultant", r"SAP\s+consultant", r"IT\s+consultant",
    r"consultant.*(?:IT|tech|SAP|cloud|data|digital|strategy|advisory)",
    # Operations / Facility / Logistics
    r"manutenz", r"manutentore", r"facility\s+specialist", r"facility\s+manager",
    r"operations?\s+(?:supervisor|manager|specialist)",
    r"supply\s+chain", r"logistic", r"warehouse", r"magazzin",
    r"operai[oa]", r"muratore", r"idraulic", r"elettricist",
    # Food / Retail / Hospitality
    r"autogrill", r"ristorazion", r"camerier", r"barista", r"cuoc[oa]",
    r"commess[oa]", r"cassier", r"addet.*vendita", r"store\s+manager",
    r"recruiting\s+day.*(?:autogrill|ristoraz|addett)", r"receptionist",
    # Finance / Legal / Consulting generico
    r"contabil", r"accountant", r"account(?:ing)?(?:\s+manager)?$",
    r"consulen.*(?:finanzi|fiscal|legale|immobili|gare|appalto)",
    r"(?:gare|appalto|bandi)", r"global\s+benefits",
    # Sales / Marketing
    r"agente.*(?:commerc|immobili|assicur)",
    r"graphic\s*design", r"marketing\s+(?:specialist|manager)",
    r"social\s+media", r"seo\b", r"copywriter",
    # employer branding è pertinente, non bloccarlo
    # Medical / Admin
    r"infermier", r"farmacist", r"segretari[oa]",
    # Misc
    r"smart\s+working.*opportunit", r"opportunit.*business",
    r"team\s+leader(?!.*(?:hr|recruit|selezione))",
    r"internship(?!.*(?:hr|recruit|selezione|talent|risorse))",
    r"stage(?!.*(?:hr|recruit|selezione|talent|risorse))",
    r"tirocinio", r"neolaureato",
    # Bandi pubblici / concorsi
    r"selezione\s+pubblica", r"concorso\s+pubblic", r"bando",
    # Organico filiale retail (non agenzie lavoro)
    r"organico\s+filiale",
])

_BLACKLIST_COMPILED = [re.compile(p, re.IGNORECASE) for p in TITLE_BLACKLIST]


def is_relevant_job(title, description=""):
    if not title:
        return False
    title_str = str(title).strip()

    # Blacklist
    for bp in _BLACKLIST_COMPILED:
        if bp.search(title_str):
            return False

    # Pattern specifici
    for rp in _RELEVANT_COMPILED:
        if rp.search(title_str):
            return True

    # Pattern agenzie dinamici
    for ap in _AGENCY_PATTERNS:
        if ap.search(title_str):
            return True

    # Keyword broad
    title_lower = title_str.lower()
    for kw in BROAD_KEYWORDS:
        if kw.lower() in title_lower:
            return True

    # Fallback: analisi descrizione
    if description:
        desc_lower = str(description).lower()
        hr_signals = [
            "talent acquisition", "recruiting", "selezione del personale",
            "ricerca e selezione", "hr business partner", "somministrazione",
            "agenzia per il lavoro", "staffing",
        ]
        title_has_hr = any(w in title_lower for w in ["hr", "human", "risorse", "personale", "staff"])
        desc_has_signal = sum(1 for s in hr_signals if s in desc_lower)
        if title_has_hr and desc_has_signal >= 2:
            return True

    return False


# === LOGGING ===

def setup_logging():
    logger = logging.getLogger("ta_monitor")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%H:%M:%S")
    sh = logging.StreamHandler(sys.stdout)
    sh.setFormatter(fmt)
    logger.addHandler(sh)
    fh = logging.FileHandler(LOG_DIR / "ta_monitor.log", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    return logger

log = setup_logging()

def load_config():
    with open(BASE_DIR / "config.yaml", "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    # Override search_terms da filters.json se presente
    if _EXTERNAL_FILTERS and "search_terms" in _EXTERNAL_FILTERS:
        cfg["search_terms"] = _EXTERNAL_FILTERS["search_terms"]
        log.info(f"  Search terms da filters.json: {len(cfg['search_terms'])} keywords")
    return cfg

# === DATABASE ===

def init_db():
    db_path = DATA_DIR / "ta_monitor.db"
    conn = sqlite3.connect(str(db_path))
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS jobs (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            company TEXT,
            location TEXT,
            channel TEXT,
            url TEXT,
            description TEXT,
            salary_min REAL,
            salary_max REAL,
            job_type TEXT,
            date_posted TEXT,
            date_scraped TEXT NOT NULL,
            search_term TEXT,
            is_new INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            total_found INTEGER,
            new_found INTEGER,
            duplicates_removed INTEGER,
            filtered_irrelevant INTEGER,
            status TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_date ON jobs(date_scraped);
        CREATE INDEX IF NOT EXISTS idx_channel ON jobs(channel);
    """)
    conn.commit()
    return conn

def job_id(title, company, url):
    raw = f"{(title or '').lower().strip()}|{(company or '').lower().strip()}|{(url or '').strip()}"
    return hashlib.md5(raw.encode()).hexdigest()

# === JOBSPY ===

def scrape_jobspy(config, hours_old):
    try:
        from jobspy import scrape_jobs
    except ImportError:
        log.error("python-jobspy non installato!")
        return pd.DataFrame()

    sites = config.get("jobspy_sites", ["indeed", "linkedin", "google"])
    location = config.get("location", "Roma, Lazio, Italia")
    proxies = config.get("proxies", []) or []
    results = []

    for term in config.get("search_terms", []):
        log.info(f"  JobSpy [{','.join(sites)}] → '{term}'")
        try:
            df = scrape_jobs(
                site_name=sites,
                search_term=term,
                google_search_term=f"{term} lavoro Roma Italia",
                location=location,
                results_wanted=config.get("jobspy_params", {}).get("results_wanted", 50),
                hours_old=hours_old,
                country_indeed="Italy",
                linkedin_fetch_description=config.get("jobspy_params", {}).get("linkedin_fetch_description", False),
                proxies=proxies if proxies else None,
            )
            if not df.empty:
                df["search_term"] = term
                results.append(df)
                log.info(f"    ✓ {len(df)} risultati")
            else:
                log.info(f"    · 0 risultati")
        except Exception as e:
            log.warning(f"    ✗ Errore: {e}")

    return pd.concat(results, ignore_index=True) if results else pd.DataFrame()

# === CUSTOM SCRAPERS ===

def scrape_custom(config):
    try:
        import requests
        from bs4 import BeautifulSoup
    except ImportError:
        log.warning("requests/bs4 mancanti, skip scraper custom")
        return []

    custom_cfg = config.get("custom_scrapers", {})
    jobs = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0",
        "Accept-Language": "it-IT,it;q=0.9",
    }
    for site, cfg in custom_cfg.items():
        if not cfg.get("enabled"):
            continue
        for term in config.get("search_terms", [])[:4]:
            url = cfg["base_url"].format(keyword=term.replace(" ", "+"))
            log.info(f"  Custom [{site}] → '{term}'")
            try:
                resp = requests.get(url, headers=headers, timeout=15)
                if resp.status_code != 200:
                    log.warning(f"    ✗ HTTP {resp.status_code}")
                    continue
                soup = BeautifulSoup(resp.text, "html.parser")
                found = _parse_job_cards(soup, site, term)
                jobs.extend(found)
                log.info(f"    ✓ {len(found)} risultati")
            except Exception as e:
                log.warning(f"    ✗ {e}")
    return jobs

def _parse_job_cards(soup, site_name, search_term):
    jobs = []
    for sel in ["article", ".job-card", ".job-item", "[class*='job']", ".search-result"]:
        cards = soup.select(sel)
        if len(cards) >= 2:
            break
    else:
        cards = []

    for card in cards[:30]:
        title_el = card.select_one("h2, h3, h4, [class*='title']")
        title = title_el.get_text(strip=True) if title_el else None
        if not title:
            continue
        company_el = card.select_one("[class*='company'], [class*='employer']")
        location_el = card.select_one("[class*='location'], [class*='city']")
        link_el = card.select_one("a[href]")
        desc = card.get_text(strip=True)[:500]

        if not is_relevant_job(title, desc):
            continue

        jobs.append({
            "title": title,
            "company": company_el.get_text(strip=True) if company_el else site_name.capitalize(),
            "location": location_el.get_text(strip=True) if location_el else "Roma",
            "channel": site_name.capitalize(),
            "url": link_el.get("href", "") if link_el else "",
            "search_term": search_term,
            "description": desc,
        })
    return jobs

# === FILTRO RILEVANZA POST-SCRAPING ===

def filter_relevant(df):
    if df.empty:
        return df, 0
    mask = df.apply(lambda row: is_relevant_job(row.get("title", ""), row.get("description", "")), axis=1)
    filtered = df[mask].reset_index(drop=True)
    removed = len(df) - len(filtered)
    log.info(f"  Filtro rilevanza: {len(df)} → {len(filtered)} (scartati {removed} non pertinenti)")
    if removed > 0:
        for _, row in df[~mask].head(10).iterrows():
            log.info(f"    ✗ Scartato: \"{row.get('title', '?')}\"")
    return filtered, removed

# === DEDUPLICAZIONE ===

def deduplicate(df, threshold=85):
    if df.empty:
        return df
    seen, keep = [], []
    for _, row in df.iterrows():
        t = str(row.get("title", "")).lower().strip()
        c = str(row.get("company", "")).lower().strip()
        dup = False
        for st, sc in seen:
            score = fuzz.ratio(t, st)
            if c and sc and fuzz.ratio(c, sc) > 80:
                score += 10
            if score >= threshold:
                dup = True
                break
        seen.append((t, c))
        keep.append(not dup)
    orig = len(df)
    result = df[keep].reset_index(drop=True)
    log.info(f"  Dedup: {orig} → {len(result)} (rimossi {orig - len(result)})")
    return result

# === STORICO ===

def find_new(conn, df):
    if df.empty:
        return df
    existing = {r[0] for r in conn.execute("SELECT id FROM jobs")}
    mask = [job_id(row.get("title",""), row.get("company",""), row.get("job_url", row.get("url",""))) not in existing
            for _, row in df.iterrows()]
    new = df[mask].reset_index(drop=True)
    log.info(f"  Nuove: {len(new)} su {len(df)} totali")
    return new

def save_to_db(conn, df):
    now = datetime.now().isoformat()
    n = 0
    for _, row in df.iterrows():
        jid = job_id(row.get("title",""), row.get("company",""), row.get("job_url", row.get("url","")))
        dp = row.get("date_posted", "")
        if pd.isna(dp) or str(dp) in ("None", "NaT", "nan", ""):
            dp = ""
        else:
            dp = str(dp)[:10]
        try:
            conn.execute(
                "INSERT OR IGNORE INTO jobs (id,title,company,location,channel,url,description,"
                "salary_min,salary_max,job_type,date_posted,date_scraped,search_term,is_new) "
                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,1)",
                (jid, row.get("title",""), row.get("company",""),
                 str(row.get("location", row.get("city","Roma"))),
                 str(row.get("site", row.get("channel",""))),
                 row.get("job_url", row.get("url","")),
                 str(row.get("description",""))[:2000],
                 row.get("min_amount"), row.get("max_amount"),
                 row.get("job_type",""), dp, now, row.get("search_term","")))
            n += 1
        except Exception as e:
            log.warning(f"  DB error: {e}")
    conn.commit()
    return n

# === REPORT EXCEL ===

def generate_excel(conn):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    today = datetime.now().strftime("%Y-%m-%d")
    path = OUTPUT_DIR / f"TA_Monitor_Roma_{today}.xlsx"
    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="4A00E0")

    ws = wb.active
    ws.title = "Nuove Posizioni"
    headers = ["Titolo", "Azienda", "Canale", "Luogo", "Tipo", "Salary Min", "Salary Max", "URL", "Data"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill = hfont, hfill
        cell.alignment = Alignment(horizontal="center")

    cutoff = (datetime.now() - timedelta(hours=168)).isoformat()
    rows = conn.execute(
        "SELECT title, company, channel, location, job_type, salary_min, salary_max, url, date_posted "
        "FROM jobs WHERE date_scraped > ? ORDER BY date_scraped DESC", (cutoff,)
    ).fetchall()

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            display = val if val and str(val) not in ("None", "nan", "NaT", "") else ""
            cell = ws.cell(row=r, column=c, value=display)
            if c == 8 and val:
                cell.font = Font(color="0066CC", underline="single")

    for c, w in enumerate([45, 30, 12, 25, 12, 12, 12, 55, 12], 1):
        ws.column_dimensions[chr(64 + c)].width = w
    if rows:
        ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Per Canale")
    for c, h in enumerate(["Canale", "Questa Scan", "Totale"], 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font, cell.fill = hfont, hfill
    for r, row in enumerate(conn.execute(
        "SELECT channel, SUM(CASE WHEN date_scraped > ? THEN 1 ELSE 0 END), COUNT(*) "
        "FROM jobs GROUP BY channel ORDER BY 2 DESC", (cutoff,)), 2):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val or "")

    ws3 = wb.create_sheet("Trend 7gg")
    for c, h in enumerate(["Data", "Posizioni"], 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font, cell.fill = hfont, hfill
    week_ago = (datetime.now() - timedelta(days=7)).isoformat()
    for r, row in enumerate(conn.execute(
        "SELECT DATE(date_scraped), COUNT(*) FROM jobs WHERE date_scraped > ? "
        "GROUP BY DATE(date_scraped) ORDER BY 1 DESC", (week_ago,)), 2):
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)

    wb.save(path)
    log.info(f"  Excel: {path}")
    return path

# === SUMMARY ===

def write_summary(output_dir, total, new, dupes, filtered, channels, profile_id="default"):
    s = {"date": datetime.now().isoformat(), "profile": profile_id,
         "total_raw": total, "filtered_irrelevant": filtered,
         "new_jobs": new, "duplicates_removed": dupes, "channels": channels}
    with open(output_dir / "summary.json", "w") as f:
        json.dump(s, f, indent=2)
    return s

# === PROFILES ===

def load_profiles():
    """Carica profiles.json se presente."""
    fpath = BASE_DIR / "profiles.json"
    if fpath.exists():
        with open(fpath, "r", encoding="utf-8") as f:
            return json.load(f)
    return None

def get_profile_filters(profile_id):
    """Restituisce i filtri per un profilo specifico."""
    profiles_data = load_profiles()
    if not profiles_data or "profiles" not in profiles_data:
        return None
    return profiles_data["profiles"].get(profile_id)

def setup_profile_dirs(profile_id):
    """Crea e restituisce le directory per un profilo."""
    if profile_id == "default" or profile_id == "hr-roma":
        # Backward compat: profilo default usa le cartelle originali
        out = OUTPUT_DIR
        db_path = DATA_DIR / "ta_monitor.db"
    else:
        out = OUTPUT_DIR / profile_id
        db_path = DATA_DIR / f"{profile_id}.db"
    out.mkdir(parents=True, exist_ok=True)
    return out, db_path

def build_config_from_profile(profile):
    """Costruisce un config dict dal profilo."""
    f = profile.get("filters", {})
    base_config = load_config()
    base_config["search_terms"] = f.get("search_terms", base_config.get("search_terms", []))
    base_config["location"] = profile.get("location", base_config.get("location", "Roma, Lazio, Italia"))
    return base_config

def build_filters_from_profile(profile):
    """Costruisce i filtri (relevant, blacklist, broad, agencies) dal profilo."""
    f = profile.get("filters", {})
    return {
        "relevant_patterns": f.get("relevant_patterns", []),
        "broad_keywords": f.get("broad_keywords", []),
        "blacklist": f.get("blacklist", []),
        "agencies": f.get("agencies", []),
    }

# === MAIN ===

def run_profile(profile_id, profile, hours_old, search_terms_override=None):
    """Esegue lo scan per un singolo profilo."""
    profile_name = profile.get("name", profile_id)
    log.info("=" * 55)
    log.info(f"TA MONITOR v4.0 — Profilo: {profile_name}")
    log.info(f"  ID: {profile_id} | Finestra: {hours_old}h | Location: {profile.get('location', '?')}")
    log.info("=" * 55)

    # Setup dirs
    out_dir, db_path = setup_profile_dirs(profile_id)

    # Build config and filters
    config = build_config_from_profile(profile)
    if search_terms_override:
        config["search_terms"] = [t.strip() for t in search_terms_override.split(",")]

    # Override global filters with profile-specific ones
    pf = build_filters_from_profile(profile)

    # Compile profile-specific patterns
    rel_patterns = [re.compile(p["pattern"] if isinstance(p, dict) else p, re.IGNORECASE) for p in pf.get("relevant_patterns", [])]
    bl_patterns = [re.compile(p["pattern"] if isinstance(p, dict) else p, re.IGNORECASE) for p in pf.get("blacklist", [])]
    broad_kw = pf.get("broad_keywords", [])
    agencies = pf.get("agencies", [])

    # Build agency patterns
    agency_compiled = []
    if agencies:
        parts = [re.escape(a).replace(r"\ ", r"\s*") for a in agencies]
        agency_re = "|".join(parts)
        roles_re = r"(?:filiale|branch|responsabil|dirett|manager|account|hr\s+consultant|recruiter|specialist)"
        agency_compiled = [
            re.compile(f"(?:{agency_re}).*{roles_re}", re.IGNORECASE),
            re.compile(f"{roles_re}.*(?:{agency_re})", re.IGNORECASE),
        ]

    def profile_is_relevant(title, description=""):
        if not title:
            return False
        title_str = title.strip()
        # Blacklist
        for bp in bl_patterns:
            if bp.search(title_str):
                return False
        # Relevant patterns
        for rp in rel_patterns:
            if rp.search(title_str):
                return True
        # Agency patterns
        for ap in agency_compiled:
            if ap.search(title_str):
                return True
        # Broad keywords fallback
        combined = f"{title_str} {description}".lower()
        return any(kw.lower() in combined for kw in broad_kw)

    # DB
    conn = sqlite3.connect(str(db_path))
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS jobs (
            id TEXT PRIMARY KEY, title TEXT NOT NULL, company TEXT,
            location TEXT, channel TEXT, url TEXT, description TEXT,
            salary_min REAL, salary_max REAL, job_type TEXT,
            date_posted TEXT, date_scraped TEXT NOT NULL,
            search_term TEXT, is_new INTEGER DEFAULT 1
        );
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT, timestamp TEXT NOT NULL,
            total_found INTEGER, new_found INTEGER, duplicates_removed INTEGER,
            filtered_irrelevant INTEGER, status TEXT
        );
        CREATE INDEX IF NOT EXISTS idx_date ON jobs(date_scraped);
    """)
    conn.commit()

    # Scrape
    log.info("▸ Fase 1: Scraping JobSpy")
    jobspy_df = scrape_jobspy(config, hours_old)

    log.info("▸ Fase 2: Scraping agenzie")
    custom_jobs = scrape_custom(config)

    log.info("▸ Fase 3: Merge")
    if custom_jobs:
        custom_df = pd.DataFrame(custom_jobs)
        all_df = pd.concat([jobspy_df, custom_df], ignore_index=True) if not jobspy_df.empty else custom_df
    else:
        all_df = jobspy_df
    total_raw = len(all_df)
    log.info(f"  Totale grezzo: {total_raw}")

    if all_df.empty:
        log.warning("Nessun risultato. Fine.")
        write_summary(out_dir, 0, 0, 0, 0, [], profile_id)
        conn.close()
        return

    log.info("▸ Fase 4: Filtro rilevanza")
    mask = all_df.apply(lambda row: profile_is_relevant(row.get("title", ""), row.get("description", "")), axis=1)
    relevant_df = all_df[mask].reset_index(drop=True)
    filtered_count = len(all_df) - len(relevant_df)
    log.info(f"  Filtro: {len(all_df)} → {len(relevant_df)} (scartati {filtered_count})")

    if relevant_df.empty:
        log.warning("Nessuna posizione rilevante.")
        write_summary(out_dir, total_raw, 0, 0, filtered_count, [], profile_id)
        conn.close()
        return

    log.info("▸ Fase 5: Deduplicazione")
    deduped = deduplicate(relevant_df)

    log.info("▸ Fase 6: Confronto storico")
    new_df = find_new(conn, deduped)

    log.info("▸ Fase 7: Salvataggio")
    save_to_db(conn, deduped)

    log.info("▸ Fase 8: Report")
    # Save output to profile dir
    _orig_output = globals().get("OUTPUT_DIR")
    today = datetime.now().strftime("%Y-%m-%d")
    generate_excel_profile(conn, out_dir, profile_name, hours_old)
    deduped.to_csv(out_dir / f"ta_jobs_{today}.csv", index=False, quoting=csv.QUOTE_NONNUMERIC)
    deduped.to_json(out_dir / f"ta_jobs_{today}.json", orient="records", force_ascii=False, indent=2)

    channels = list(deduped["site"].unique()) if "site" in deduped.columns else []
    write_summary(out_dir, total_raw, len(new_df), len(relevant_df) - len(deduped), filtered_count, channels, profile_id)

    try:
        conn.execute("ALTER TABLE runs ADD COLUMN filtered_irrelevant INTEGER")
    except sqlite3.OperationalError:
        pass
    conn.execute(
        "INSERT INTO runs (timestamp,total_found,new_found,duplicates_removed,filtered_irrelevant,status) VALUES (?,?,?,?,?,?)",
        (datetime.now().isoformat(), total_raw, len(new_df), len(relevant_df) - len(deduped), filtered_count, "success"))
    conn.commit()
    conn.close()

    log.info(f"✓ [{profile_name}] Grezzo: {total_raw} | Scartati: {filtered_count} | Rilevanti: {len(deduped)} | Nuove: {len(new_df)}")
    log.info("=" * 55)


def generate_excel_profile(conn, out_dir, profile_name, hours_old):
    """Genera Excel nella directory del profilo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    today = datetime.now().strftime("%Y-%m-%d")
    path = out_dir / f"TA_Monitor_{today}.xlsx"
    wb = Workbook()
    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="4A00E0")

    ws = wb.active
    ws.title = "Posizioni"
    headers = ["Titolo", "Azienda", "Canale", "Luogo", "Tipo", "Salary Min", "Salary Max", "URL", "Data"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill = hfont, hfill
        cell.alignment = Alignment(horizontal="center")

    cutoff = (datetime.now() - timedelta(hours=hours_old)).isoformat()
    rows = conn.execute(
        "SELECT title, company, channel, location, job_type, salary_min, salary_max, url, date_posted "
        "FROM jobs WHERE date_scraped > ? ORDER BY date_scraped DESC", (cutoff,)
    ).fetchall()

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            display = val if val and str(val) not in ("None", "nan", "NaT", "") else ""
            cell = ws.cell(row=r, column=c, value=display)
            if c == 8 and val:
                cell.font = Font(color="0066CC", underline="single")

    for c, w in enumerate([45, 30, 12, 25, 12, 12, 12, 55, 12], 1):
        ws.column_dimensions[chr(64 + c)].width = w
    if rows:
        ws.auto_filter.ref = f"A1:I{len(rows) + 1}"
    ws.freeze_panes = "A2"
    wb.save(path)
    log.info(f"  Excel: {path}")
    return path


def main():
    parser = argparse.ArgumentParser(description="TA Monitor v4.0 — Multi-Profile")
    parser.add_argument("--hours-old", type=int, default=168)
    parser.add_argument("--search-terms", type=str, default="")
    parser.add_argument("--profile", type=str, default="all",
                        help="Profile ID to run, or 'all' for all auto_scan profiles")
    args = parser.parse_args()

    profiles_data = load_profiles()

    if not profiles_data or "profiles" not in profiles_data:
        # Fallback: run legacy mode with filters.json
        log.info("Nessun profiles.json trovato, modalità legacy")
        config = load_config()
        if args.search_terms:
            config["search_terms"] = [t.strip() for t in args.search_terms.split(",")]
        # Wrap in a fake profile
        legacy_profile = {
            "name": "Legacy",
            "location": config.get("location", "Roma, Lazio, Italia"),
            "filters": _EXTERNAL_FILTERS or {},
        }
        run_profile("default", legacy_profile, args.hours_old, args.search_terms or None)
        return

    profiles = profiles_data["profiles"]

    if args.profile == "all":
        # Run all auto_scan profiles
        active = {pid: p for pid, p in profiles.items() if p.get("auto_scan", False)}
        if not active:
            log.warning("Nessun profilo con auto_scan attivo")
            return
        log.info(f"Profili auto_scan: {', '.join(active.keys())}")
        for pid, profile in active.items():
            hours = profile.get("hours_old", args.hours_old)
            try:
                run_profile(pid, profile, hours)
            except Exception as e:
                log.error(f"Errore profilo {pid}: {e}")
    else:
        # Run specific profile
        if args.profile not in profiles:
            log.error(f"Profilo '{args.profile}' non trovato. Disponibili: {', '.join(profiles.keys())}")
            sys.exit(1)
        profile = profiles[args.profile]
        hours = profile.get("hours_old", args.hours_old)
        run_profile(args.profile, profile, hours, args.search_terms or None)


if __name__ == "__main__":
    main()